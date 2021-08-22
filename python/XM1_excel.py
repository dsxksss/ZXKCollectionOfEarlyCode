#处理excel表格小项目
import openpyxl as xl  #从openpyxl基本库中导入操作excel的类
from openpyxl.chart import BarChart, Reference

#以下内容可以定义一个函数方法,方便修改fileName
dx = xl.load_workbook('bigdata.xlsx')  #要处理的excel文件名
sheet = dx['Sheet1']  #表示处理这个文件的第一个表,S大写
#cell = sheet['b1']  #通过单元格”坐标名“存放内容
#竖一条是行,横一条是列!!!
#cell = sheet.cell(1, 2)  #通过单元格“坐标”,标识内容,C1:列,C2:行
#print(cell.value)  #打印其value
#如果你修改了,或者动了一下某个单元格的数据,都会影响打印行列max
#print(sheet.max_row)  #打印次标签的行数,这里可以利用循环,获取某行数据

for sum in range(2, sheet.max_row + 1):
    #从2开始(包括2自己)到max_row之间的数(不包括max_row自己,所以要加1)
    #print(sum)
    cell = sheet.cell(sum, 2)  #一开始这里指表格第二列第二行依此叠加
    print(cell.value)  #查看这个单元格的value
    age_new1 = cell.value + 10  #每个年龄加10,结果存放age_new1
    age_new2 = sheet.cell(sum, 4)  #在创建表格第四行的位置
    age_new2.value = age_new1  #将结果放到新创建的位置

#预设好一个表格的数据、大小、边框(盲猜...)
value_map = Reference(sheet,
                      min_row=2,
                      max_row=sheet.max_row,
                      min_col=4,
                      max_col=4)
chart = BarChart()  #创建表格控件
chart.add_data(value_map)  #将表格数据绑定至此控件
sheet.add_chart(chart, 'e2')  #将此控件在e2位置开始展示

#表格保存的路径,可以是当前的也可以是新的,这里是防止意外创建的新文件
dx.save('bigdate2.xlsx')
