#bat.1.0.15
# 于2021.9.21.18:43结束
#            总·完成功能:
# 1、实现从外部的excel表格读取信息
# 2、将读取好的信息存入到一个二维列表里
# 3、将二维列表里的数据实现输入学生名字，即可查看此人全部信息
# 4、输入学生电话号码查询功能
# 5、将此项目设置为类结构化
# 6、显示读取数据的类型(非表格对其化)
# 7、拥有更多兼容性,可以处理其他表格而不局限于目前此表格

#TODO:orFIXME:
# {
# 1、增加额外功能
# 2、将此项目会发生异常地方加上解决方法(已解决一处,,,)
#、、、、、、
# }


class XM():
    def __init__(self):
        pass

    #main
    def main(self,
             excel_name,
             excel_sheets="Sheet1",
             to_whatuser_data=1,
             user_data_size=5):
        import openpyxl as xl  #excel表格处理库
        import pretty_errors  #美化errors库

        student_data_list = []  # 存放学生信息
        excel_sheet = xl.load_workbook(excel_name)  # 要处理的表格路径名字
        sheet = excel_sheet[excel_sheets]  # 表示要读取此文件的第一页表格,这里的Sheet1开头s必须大写
        man_size = to_whatuser_data  #开始从第x个位置读取数据

        for sum_one in range(sheet.max_row):  #总共录入的学生人数
            for sum in range(2, sheet.max_column):  #一个学生的总信息个数
                my_cell = sheet.cell(man_size, sum)  # 2表示表格行,sum表示表格列
                student_data_list.append(
                    my_cell.value)  #读到的数据存入到name_list里(还是一维列表DATA)
            man_size += 1  #读取完一个人信息然后自加1,继续读取下一个人信息

        student_data_temporary = []  #暂存一个人的全部信息
        student_data_biglist = []  #最后的结果,是一个二维列表

        #may TODO:此处可以利用切片实现，但我暂时不知会不会影响系统开销，所以暂时不利用
        for sum in student_data_list:  #读取五个数据存储一次
            student_data_temporary.append(sum)
            if len(student_data_temporary
                   ) >= user_data_size:  #如果暂存数据达到五个(或其他任意值)之后算一个人的信息
                student_data_biglist.append(student_data_temporary)
                student_data_temporary = []  #达到要求之后,清空暂存数据,继续接受至无数据

        print("\n\n\n\n")
        print(student_data_list)

        #按一个学生的全部信息显示(data已归纳)
        # for sum in range(len(student_data_biglist)):
        #     print(student_data_biglist[sum])

        print("人数总共读取", "[\033[1;32m",
              len(student_data_biglist) - 1, "\033[0m]", "个")  #彩色显示读取的人数
        print("数据总共读取", "[\033[1;32m",
              len(student_data_list) - 5, "\033[0m]",
              "条")  #彩色显示读取的总数据(data未进行归纳)
        return student_data_biglist
#查询函数(也可以使用index函数查找内容下标,这里没使用是因为不会用index查找二维列表😅)

    def find_name(self, data_input, list_Tofind):  #C1:外界输入,C2:要进行搜索的data
        for sum in range(len(list_Tofind)):
            if (list_Tofind[sum][0] == data_input) or str(
                    list_Tofind[sum][4]) == data_input:  #尽量别出现除str类型的其他强转类型比较
                print("\033[1;31m已查询到:\033[0m\n\033[1;32m", list_Tofind[0],
                      "\n", list_Tofind[sum], "\033[0m")
                return True  #查询到则返回True
        return False  #为查询到则返回False


#(程序开始主入口)
待处理表格 = "database.xlsx"
dx = XM()  #创建类对象
#填写要读取的excel表格and要处理此表格的表页and开始从第x个位置读取数据and一个人拥有信息的总个数
#(表页默认只处理第一页,如要更改可在代码第23处改动)
data = dx.main(excel_name=待处理表格)
#填写要查找的内容and要查找的数据
try:
    if dx.find_name(data_input=input("请输入要查询的学生姓名或手机联系号码 >"),
                    list_Tofind=data):
        pass
    else:
        print("\033[1;31m数据库中暂无此人!!!\033[0m")
except ValueError:  #捕捉异常输入
    print("\033[1;31mError发生错误![您输入的值不符合搜索规范或含有风险代码,请重开程序再使用]\033[0m")